import json
import pandas as pd
import re
from pathlib import Path
from loguru import logger
from openpyxl.styles import Alignment, PatternFill, Font

# Configuration
REGISTRY_PATH = Path("data/fsm_registry.json")
OUTPUT_EXCEL = Path("data/Matrice_Conformite_RFP.xlsx")

class ExcelGenerator:
    """Générateur de précision pour la matrice de conformité (v13)."""

    def __init__(self, registry_path: Path):
        self.registry_path = registry_path
        self.categories = {"MUST": [], "SHOULD": [], "COULD": [], "WONT": []}
        self.rejected_noise = []
        self.seen_hashes = set()

    def _is_noise(self, entry: dict) -> bool:
        """Détecteur de bruit haute précision."""
        sujet = str(entry.get("subject", "")).strip()
        action = str(entry.get("action", "")).strip()
        objet = str(entry.get("target_object", "")).strip()
        source = str(entry.get("source_quote", "")).strip()
        
        # 1. Critères de vide
        if any(not val or val.lower() in ["null", "none", "aucun", "n/a", "unknown"] for val in [sujet, action, objet]):
            return True
            
        # 2. Critères de longueur (Anti-bruit atomique)
        if len(sujet) < 3 or len(objet) < 3:
            return True

        # 3. Patterns de structure documentaires
        noise_patterns = [
            r"table of", r"page \d+", r"figure \d+", r"appendix", r"end of",
            r"confidential", r"iss\.", r"applicable version", r"certified by",
            r"section:", r"\|", r"---", r"___"
        ]
        if any(re.search(p, source, re.IGNORECASE) for p in noise_patterns):
            return True

        # 4. Anti-doublon sémantique local (basé sur la citation)
        quote_hash = hash(source.lower().strip())
        if quote_hash in self.seen_hashes:
            return True
        self.seen_hashes.add(quote_hash)

        return False

    def _get_moscow(self, action: str, source: str) -> str:
        """Tri MoSCoW avec gestion des interdictions."""
        text = (action + " " + source).lower()
        
        if any(w in text for w in ["cannot", "must not", "shall not", "interdit", "ne doit pas"]):
            return "WONT"
        if any(w in text for w in ["must", "shall", "doit", "required", "obligatoire"]):
            return "MUST"
        if any(w in text for w in ["should", "devrait", "recommended", "recommandé"]):
            return "SHOULD"
        if any(w in text for w in ["can", "could", "may", "peut", "might", "possible"]):
            return "COULD"
        return "WONT"

    def load_and_process(self):
        if not self.registry_path.exists():
            logger.error("❌ Registre introuvable.")
            return

        with open(self.registry_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        for entry in data:
            formatted = {
                "ID Officiel": entry.get("metadata", {}).get("official_id", "N/A"),
                "Sujet": entry.get("subject", ""),
                "Action": entry.get("action", ""),
                "Objet": entry.get("target_object", ""),
                "Citation Source": entry.get("source_quote", ""),
                "Page": entry.get("metadata", {}).get("page", "?"),
                "Qualité (%)": 100 - int(entry.get("ambiguity_score", 0)),
                "Statut": "À vérifier"
            }

            if self._is_noise(entry):
                formatted["Raison Rejet"] = "Bruit ou Doublon"
                self.rejected_noise.append(formatted)
            else:
                cat = self._get_moscow(entry.get("action", ""), entry.get("source_quote", ""))
                self.categories[cat].append(formatted)

    def generate(self):
        logger.info("🚀 Génération de la Matrice Haute Fidélité...")
        
        with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
            # Onglet Synthèse
            summary = pd.DataFrame([
                {"Catégorie": k, "Nombre": len(v)} for k, v in self.categories.items()
            ] + [{"Catégorie": "REJETÉS", "Nombre": len(self.rejected_noise)}])
            summary.to_excel(writer, sheet_name="Synthèse", index=False)

            # Onglets MoSCoW
            for cat, items in self.categories.items():
                df = pd.DataFrame(items)
                df.to_excel(writer, sheet_name=cat, index=False)

            # Onglet Bruit
            pd.DataFrame(self.rejected_noise).to_excel(writer, sheet_name="Rejetés (Audit)", index=False)

            # Post-traitement visuel
            self._apply_styling(writer)

        logger.success(f"✨ Matrice v13 générée : {len(self.seen_hashes)} exigences uniques retenues.")

    def _apply_styling(self, writer):
        workbook = writer.book
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)

        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            ws.freeze_panes = "A2"
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            for i, col in enumerate(ws.columns):
                col_letter = col[0].column_letter
                if sheet != "Synthèse":
                    width = 20
                    if i == 4: width = 80 # Citation
                    if i == 3: width = 40 # Objet
                    ws.column_dimensions[col_letter].width = width
                    
                    if width >= 40:
                        for cell in col:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")

if __name__ == "__main__":
    gen = ExcelGenerator(REGISTRY_PATH)
    gen.load_and_process()
    gen.generate()
